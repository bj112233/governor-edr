"""
Data utilities — CSV charting, Excel integrity, file integrity checks.
"""

import json
import os
from pathlib import Path


def chart_csv(
    path: str,
    x_col: str,
    y_cols: str,
    kind: str = "line",
    title: str = "",
    output: str = "",
):
    """F3: Render chart from CSV/XLSX using matplotlib. Returns path to saved PNG."""
    try:
        import matplotlib
        import pandas as pd

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return "❌ matplotlib/pandas not installed."
    ext = Path(path).suffix.lower()
    try:
        df = pd.read_csv(path) if ext == ".csv" else pd.read_excel(path)
    except Exception as e:
        return f"❌ Failed to read {path}: {e}"
    cols = [c.strip() for c in y_cols.split(",")] if y_cols else []
    cols = [c for c in cols if c in df.columns]
    if not cols:
        return f"❌ No valid Y columns. Available: {list(df.columns)}"
    if x_col and x_col not in df.columns:
        return f"❌ X column '{x_col}' not found. Available: {list(df.columns)}"

    fig, ax = plt.subplots(figsize=(10, 6))
    if kind == "bar":
        df.plot(x=x_col, y=cols, kind="bar", ax=ax)
    elif kind == "scatter":
        for c in cols:
            ax.scatter(df[x_col] if x_col else df.index, df[c], label=c)
        ax.legend()
    else:  # line
        df.plot(x=x_col, y=cols, kind="line", ax=ax, marker="o")
    ax.set_title(title or f"{kind.capitalize()} chart of {','.join(cols)}")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    if not output:
        output = str(Path(path).with_suffix(".png"))
    plt.savefig(output, dpi=120)
    plt.close(fig)
    return f"✅ Chart saved: {output}\n[FILE_EXPORT: {output}]\nShape: {df.shape}\nKind: {kind}, Y={cols}"


_FORMULA_ERRORS = ("#REF!", "#NAME?", "#DIV/0!", "#N/A", "#VALUE!")


def _scan_row_for_issues(row) -> tuple[int, bool, list[str]]:
    """Scan one row — returns (formula_errors, has_empty, row_vals)."""
    formula_errors = 0
    has_empty = False
    row_vals: list[str] = []
    for cell in row:
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            if any(err in v.upper() for err in _FORMULA_ERRORS):
                formula_errors += 1
        if v is None:
            has_empty = True
        row_vals.append(str(v) if v is not None else "")
    return formula_errors, has_empty, row_vals


def _evaluate_sheet(ws) -> tuple[list[str], int]:
    """Evaluate one worksheet — returns (issue_descriptions, issue_count)."""
    rows, cols = ws.max_row, ws.max_column
    formula_errors = 0
    empty_count = 0
    seen_rows: dict[tuple, int] = {}
    duplicates = 0

    for row in ws.iter_rows(min_row=1, max_row=min(rows, 10000)):
        errs, has_empty, row_vals = _scan_row_for_issues(row)
        formula_errors += errs
        if has_empty:
            empty_count += 1
        row_key = tuple(row_vals)
        if any(row_vals) and row_key in seen_rows:
            duplicates += 1
        seen_rows[row_key] = seen_rows.get(row_key, 0) + 1

    issues: list[str] = []
    if formula_errors:
        issues.append(f"{formula_errors} broken formula(s)")
    if duplicates:
        issues.append(f"{duplicates} duplicate row(s)")
    empty_pct = 100 * empty_count / max(rows * cols, 1)
    if empty_pct > 30:
        issues.append(f"{empty_pct:.0f}% empty cells")
    return issues, len(issues)


def _format_sheet_section(sheet_name: str, ws, issues: list[str]) -> list[str]:
    """Format one sheet's report section."""
    rows, cols = ws.max_row, ws.max_column
    lines = [f"## {sheet_name}", f"- Dimensions: {rows} rows x {cols} cols"]
    if issues:
        lines.extend(f"- ⚠️ {i}" for i in issues)
    else:
        lines.append("- ✅ No issues detected")
    lines.append("")
    return lines


def xlsx_integrity(path: str) -> str:
    """F4: Validate Excel workbook for broken formulas, empty cells, duplicates."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "❌ openpyxl not installed."
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        return f"❌ Failed to open: {e}"

    out = [f"# Excel Integrity Report: {Path(path).name}\n"]
    total_issues = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        issues, count = _evaluate_sheet(ws)
        total_issues += count
        out.extend(_format_sheet_section(sheet_name, ws, issues))

    out.append(f"\n**Summary:** {total_issues} issue(s) across {len(wb.sheetnames)} sheet(s).")
    return "\n".join(out)


def file_integrity_check(path: str, ext: str) -> str:
    """General file integrity check for non-Excel files."""
    p = Path(path)
    if not p.exists():
        return f"❌ File not found: {path}"
    size = p.stat().st_size
    out = [
        f"# 🔍 File Check: {p.name}\n",
        f"- **Size:** {size:,} bytes",
        f"- **Type:** {ext or '(no ext)'}",
    ]
    if ext == ".pdf":
        encrypted = False
        try:
            import pypdf

            with open(path, "rb") as f:
                reader = pypdf.PdfReader(f)
                encrypted = reader.is_encrypted
        except Exception:
            pass  # pypdf may not be installed or file may be corrupt

        try:
            import pdfplumber

            with pdfplumber.open(path) as pdf:
                pages = len(pdf.pages)
                out.append(f"- **Pages:** {pages}")
                out.append(f"- **Encrypted:** {encrypted}")
                out.append("- ✅ PDF is valid and readable")
        except Exception as e:
            msg = str(e).lower()
            if "password" in msg or "encrypted" in msg or encrypted:
                out.append(f"- ⚠️ PDF מוצפן או מוגן בסיסמה: {e}")
            else:
                out.append(f"- ❌ PDF read error: {e}")
    elif ext in (".docx",):
        try:
            from docx import Document

            doc = Document(path)
            paragraphs = len(doc.paragraphs)
            out.append(f"- **Paragraphs:** {paragraphs}")
            out.append("- ✅ DOCX is valid and readable")
        except Exception as e:
            out.append(f"- ❌ DOCX read error: {e}")
    elif ext in (".json",):
        try:
            import json as _json

            data = _json.loads(p.read_text(encoding="utf-8", errors="replace"))
            kind = type(data).__name__
            length = len(data) if isinstance(data, (list, dict)) else "—"
            out.append(f"- **Root type:** {kind}, length: {length}")
            out.append("- ✅ JSON is valid")
        except _json.JSONDecodeError as e:
            out.append(f"- ❌ JSON parse error: {e}")
    elif ext in (".csv",):
        try:
            import csv as _csv

            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                reader = _csv.reader(f)
                rows = sum(1 for _ in reader)
            out.append(f"- **Rows (approx):** {rows}")
            out.append("- ✅ CSV is readable")
        except Exception as e:
            out.append(f"- ❌ CSV read error: {e}")
    elif ext in (".txt", ".md"):
        try:
            text = p.read_text(encoding="utf-8", errors="replace")
            lines = text.count("\n")
            words = len(text.split())
            out.append(f"- **Lines:** {lines:,} | **Words:** {words:,}")
            out.append("- ✅ Text file is readable")
        except Exception as e:
            out.append(f"- ❌ Read error: {e}")
    else:
        try:
            p.read_bytes()
            out.append("- ✅ File is readable (binary)")
        except Exception as e:
            out.append(f"- ❌ Read error: {e}")
    return "\n".join(out)


def _is_valid_luhn(card_number: str) -> bool:
    """Deterministic Luhn check (mod 10) for credit-card-shaped digit strings.

    Strips non-digits before validation. Returns False for lengths outside
    13..19 (industry standard PAN range).
    """
    digits = [int(c) for c in card_number if c.isdigit()]
    if not 13 <= len(digits) <= 19:
        return False
    total = 0
    # Iterate from rightmost digit; double every second.
    for i, d in enumerate(reversed(digits)):
        if i % 2 == 1:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return total % 10 == 0


def _is_valid_israeli_id(id_number: str) -> bool:
    """Israeli Teudat Zehut check-digit validator.

    Algorithm (per Israeli Ministry of Interior spec):
      1. Strip non-digits; left-pad with zeros to exactly 9 digits.
      2. Reject if length > 9 or no digits at all.
      3. Multiply each digit by 1 or 2 alternately (positions 0,2,4,6,8 → ×1;
         positions 1,3,5,7 → ×2).
      4. If a product > 9, sum its digits (equivalent to subtracting 9).
      5. Valid iff (sum % 10) == 0.
    """
    digits_only = "".join(c for c in id_number if c.isdigit())
    if not digits_only or len(digits_only) > 9:
        return False
    padded = digits_only.zfill(9)
    total = 0
    for i, ch in enumerate(padded):
        d = int(ch) * (1 if i % 2 == 0 else 2)
        if d > 9:
            d -= 9
        total += d
    return total % 10 == 0
